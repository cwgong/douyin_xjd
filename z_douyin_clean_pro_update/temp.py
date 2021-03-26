#coding=utf8
import pymysql
import traceback
import os
import xlrd
import pandas as pd
import openpyxl
from pyhive import hive
import re
from pymongo import MongoClient
from pymongo import InsertOne
import time

def get_data():
    data_list = []
    with open("updata_product_data/check_utest_brand.txt", "r", encoding="utf-8") as f1:
        for line in f1:
            if len(line.strip()) == 0:continue
            line_list = line.strip().split("\t")
            if len(line_list) != 2:continue
            brand_name, num = line_list
            data_list.append(brand_name)
    return data_list

def get_data_v1():
    data_list = []
    with open("updata_product_data/check_utest_brand.txt", "r", encoding="utf-8") as f1:
        for line in f1:
            if len(line.strip()) == 0:continue
            data_list.append(line.strip())
    return data_list

def get_data_v2():
    data_list = []
    with open("./temp_1.txt","r",encoding="utf-8") as f1:
        for line in f1:
            if len(line.strip()) == 0:continue
            data_list.append(line.strip())
    return data_list

def get_data_v3():
    data_list = []
    with open("./temp_2.txt","r",encoding="utf-8") as f1:
        for line in f1:
            if len(line.strip()) == 0:continue
            line_list = line.strip().split("\t")
            if len(line_list) != 2:continue
            brand_name,num = line_list
            data_list.append(brand_name.strip())
    return data_list

def mysql_opt_v1():
    try:
        db = pymysql.connect(host='rm-8vb1dlcr1x89lo03z.mysql.zhangbei.rds.aliyuncs.com', \
                             port=3306, user='edw', \
                             passwd='0Zwafqams%GDL3BU', db='mysql', \
                             charset='utf8')

        cursor = db.cursor()
        data_list = []
        brand_name_list = get_data_v2()
        print(len(brand_name_list))
        for brand_name in brand_name_list:
            #数码and first_type_id <> 100006 and first_type_id <> 100001
            # sql_str = 'select product_id,product_name,brand_name, first_type_name ,brand_match_result from research_online.live_product_maintain_for_brand where (first_type_id <> 100008 and first_type_id <> 100001 and first_type_id <> 100006 and first_type_id <> 100009 and brand_name = "%s" and (brand_match_result = 1 or brand_match_result = 2))' %(brand_name)
            sql_str = 'select product_id,product_name,brand_name, first_type_name ,brand_match_result from research_online.live_product_maintain_for_brand where (first_type_id <> 100014 and first_type_id <> 100033 and brand_name = "%s" and (brand_match_result = 1 or brand_match_result = 2))' % (
                brand_name)
            # sql_str = 'select product_id,product_name,brand_name, first_type_name ,brand_match_result from research_online.live_product_maintain_for_brand where (first_type_id <> 100025 and brand_name = "%s" and brand_match_result = 0)' %(brand_name)
            #生鲜
            # sql_str = 'select product_id,product_name,brand_name, first_type_name ,brand_match_result from research_online.live_product_maintain_for_brand where (first_type_id <> 100002 and brand_name = "%s" and (brand_match_result = 1 or brand_match_result = 2))' % (
            #     brand_name)
            # sql_str = 'select product_id,product_name,brand_name, first_type_name ,brand_match_result from research_online.live_product_maintain_for_brand where (first_type_id <> 100002 and brand_name = "%s" and brand_match_result = 0)' % (
            #     brand_name)
            # 家用电器
            # sql_str = 'select product_id,product_name,brand_name, first_type_name ,brand_match_result from research_online.live_product_maintain_for_brand where (first_type_id <> 100031 and first_type_id <> 100025 and brand_name = "%s" and (brand_match_result = 1 or brand_match_result = 2))' % (
            #     brand_name)
            # sql_str = 'select product_id,product_name,brand_name, first_type_name ,brand_match_result from research_online.live_product_maintain_for_brand where (first_type_id <> 100031 and first_type_id <> 100025 and brand_name = "%s" and brand_match_result = 0)' % (
            #     brand_name)
            cursor.execute(sql_str)

        # cursor.execute("select product_id, brand_id, brand_name from research_online.live_product_maintain_for_brand q where brand_match_result > 0")
            data = cursor.fetchall()
            data_list.append(data)
        cursor.close()
        db.close()
        return data_list
    except:
        print(traceback.format_exc())

def mysql_opt():
    try:
        db = pymysql.connect(host='rm-8vb1dlcr1x89lo03z.mysql.zhangbei.rds.aliyuncs.com', \
                             port=3306, user='edw', \
                             passwd='0Zwafqams%GDL3BU', db='mysql', \
                             charset='utf8')

        cursor = db.cursor()
        data_list = []
        # industry_name_list = ['家用电器','数码','生鲜']
        industry_name_dict = {'生鲜':100002,'数码':100027,'家用电器':100031}
        print(len(industry_name_dict))
        for industry_name in industry_name_dict.keys():
            sql_str = 'select product_id,product_name,brand_name, first_type_name from research_online.live_product_maintain_for_brand where (first_type_id <> %s and industry_name = "%s")' %(industry_name_dict[industry_name], industry_name)
            print(sql_str)
            cursor.execute(sql_str)

        # cursor.execute("select product_id, brand_id, brand_name from research_online.live_product_maintain_for_brand q where brand_match_result > 0")
            data = cursor.fetchall()
            data_list.append(data)
        cursor.close()
        db.close()
        return data_list
    except:
        print(traceback.format_exc())

def mysql_opt_v2():
    try:
        db = pymysql.connect(host='rm-8vb1dlcr1x89lo03z.mysql.zhangbei.rds.aliyuncs.com', \
                             port=3306, user='edw', \
                             passwd='0Zwafqams%GDL3BU', db='mysql', \
                             charset='utf8')

        cursor = db.cursor()
        data_list = []
        industry_name_list = ['家用电器','数码','生鲜']
        industry_name_dict = {'生鲜':100002,'数码':100027,'家用电器':100031}
        print(len(industry_name_dict))
        for industry_name in industry_name_list:
            sql_str = 'select product_id,product_name,brand_name, first_type_name from research_online.live_product_maintain_for_brand where (first_type_id <> %s and industry_name = "%s")' %(industry_name_dict[industry_name], industry_name)
            print(sql_str)
            cursor.execute(sql_str)

        # cursor.execute("select product_id, brand_id, brand_name from research_online.live_product_maintain_for_brand q where brand_match_result > 0")
            data = cursor.fetchall()
            data_list.append(data)
        cursor.close()
        db.close()
        return data_list
    except:
        print(traceback.format_exc())

def stat_brand_num():
    try:
        db = pymysql.connect(host='rm-8vb1dlcr1x89lo03z.mysql.zhangbei.rds.aliyuncs.com', \
                             port=3306, user='edw', \
                             passwd='0Zwafqams%GDL3BU', db='mysql', \
                             charset='utf8')

        cursor = db.cursor()
        # industry_name_list = ['家用电器','数码','生鲜']
        # sql_str = 'select destinct product_id,product_name,brand_name, first_type_name, brand_match_result from research_online.live_product_maintain_for_brand where brand_match_result = 0'
        sql_str = 'SELECT brand_name,count(1) AS num from research_online.live_product_maintain_for_brand where brand_match_result = 0 and product_source = "抖音-自营" GROUP BY brand_name ORDER BY num DESC'
        # sql_str = 'select brand_name,brand_id, count(1) as num from research_online.live_product_maintain_for_brand where brand_match_result != 0 and product_source = "拼多多" GROUP BY brand_name ORDER BY num desc'
        # print(sql_str)
        cursor.execute(sql_str)

        # cursor.execute("select product_id, brand_id, brand_name from research_online.live_product_maintain_for_brand q where brand_match_result > 0")
        data = cursor.fetchall()
        cursor.close()
        db.close()
        return data
    except:
        print(traceback.format_exc())

def get_online_brand_v1():
    data_list = get_data_v1()
    brand_list = []
    with open("./temp.txt","r",encoding="utf-8") as f1:
        for line in f1:
            if len(line) == 0:continue
            line_list = line.strip().split("\t")
            if len(line_list) != 2:continue
            brand_name ,num = line_list
            if brand_name in data_list:
                brand_list.append(line)
            else:
                continue
    print(len(brand_list))
    print("".join(brand_list))

def get_online_brand():
    data_list = get_data()
    print(len(data_list))
    brand_list = []
    with open("./temp.txt","r",encoding="utf-8") as f1:
        for line in f1:
            if len(line) == 0:continue
            line_list = line.strip().split("\t")
            if len(line_list) != 2:continue
            brand_name ,num = line_list
            if brand_name not in data_list:
                brand_list.append(line)
            else:
                continue
    print("".join(brand_list))

# def deal_file():
#     file_dir = 'C:/Users/Cwgong/Desktop/data_2/'
#     pass



def get_product():
    data_list = []
    brand_list = get_data_v3()
    try:
        db = pymysql.connect(host='rm-8vb1dlcr1x89lo03z.mysql.zhangbei.rds.aliyuncs.com', \
                             port=3306, user='edw', \
                             passwd='0Zwafqams%GDL3BU', db='mysql', \
                             charset='utf8')

        cursor = db.cursor()
        # industry_name_list = ['家用电器','数码','生鲜']
        for brand in brand_list:
            # sql_str = 'select destinct product_id,product_name,brand_name, first_type_name, brand_match_result from research_online.live_product_maintain_for_brand where brand_match_result = 0'
            sql_str = 'select product_id,product_name,brand_name, first_type_name ,brand_match_result from research_online.live_product_maintain_for_brand where product_source = "抖音-自营" and brand_match_result = 0 and brand_name = "%s"' % (brand)
            # sql_str = 'select brand_name, count(1) as num from research_online.live_product_maintain_for_brand where brand_match_result = 0 and product_source = "拼多多" GROUP BY brand_name ORDER BY num desc'
            print(sql_str)
            cursor.execute(sql_str)

            # cursor.execute("select product_id, brand_id, brand_name from research_online.live_product_maintain_for_brand q where brand_match_result > 0")
            data = cursor.fetchall()
            data_list.append(data)
        cursor.close()
        db.close()
        return data_list
    except:
        print(traceback.format_exc())

def compare_brand():
    data_list1 = []
    data_list2 = []
    with open("./temp_1.txt","r",encoding="utf-8") as f1:
        for line in f1:
            if len(line.strip()) == 0:continue
            line_list = line.strip().split("\t")
            if len(line_list) != 3:continue
            data_list1.append(line)
    with open("./temp_2.txt","r",encoding="utf-8") as f2:
        for line in f2:
            if len(line.strip()) == 0:continue
            line_list = line.strip().split("\t")
            if len(line_list) != 3:continue
            brand_name,brand_id,num = line_list
            data_list2.append(brand_id)
    print(len(data_list1))
    print(len(data_list2))
    for line in data_list1:
        line_list = line.strip().split("\t")
        brand_id, brand_name, gmv = line_list
        if brand_id not in data_list2:
            print(line.strip())
        else:
            pass


def collect_data():
    brand_name = 0
    product_num = 0
    legal_product_num = 0
    product_list = []
    file_dir = "C:/Users/Cwgong/Desktop/反馈文档/"
    for root, dirs, files in os.walk(file_dir):
        for file in files:
            # 获取文件所属目录
            # print(root)
            # 获取文件路径
            # with open(os.path.join(root, file),"r",encoding="utf-8") as f1:
            #     idx_ = 0
            #     for line in f1:
            #         if idx_ == 0:continue
            #         idx_ += 1
            #         if len(line.strip()) == 0:continue
            #         line_list = line.strip().split("\t")
            #         if len(line_list) !=
            # workbook = xlrd.open_workbook(os.path.join(root, file))
            wb = openpyxl.load_workbook(os.path.join(root, file))
            # sheets = workbook.sheet_names()
            sheets = wb.sheetnames
            # print(sheets)
            brand_name = brand_name + len(sheets)
            for i in range(0,len(sheets)):
            #     df = pd.read_excel(os.path.join(root, file),sheet_name = i,skiprows=1, index = False,encoding = "utf-8")
            #     df28.append(df)
            # print(df28)
                sheet = wb[sheets[i]]

                # print('\n\n第' + str(i + 1) + '个sheet: ' + sheet.title + '->>>')
                for r in range(1, sheet.max_row + 1):
                    if r == 1:
                        # print('\n' + ''.join(
                        #     [str(sheet.cell(row=r, column=c).value).ljust(17) for c in range(1, sheet.max_column + 1)]))
                        continue
                    else:
                        if str(sheet.cell(row=r, column=1).value) == "1":
                            legal_product_num = legal_product_num + 1
                            product_num = product_num + 1
                            product_str = ''.join(
                                [str(sheet.cell(row=r, column=c).value) + "\t" for c in range(1, sheet.max_column + 1)])
                            print(product_str)
                            product_list.append(product_str)
                        else:
                            product_num = product_num + 1

    print("品牌数目:" + str(brand_name))
    print("审核总商品数:" + str(product_num))
    print("审核结果为小家电商品数:" + str(legal_product_num))
    print("审核结果为小家电商品数/审核总商品数:" + str(legal_product_num/product_num))

    # for product in product_list:
    #     if len(product) == 0:continue
    #     product_list = product.strip().split("\t")
    #     if len(product_list) != 8:continue
    #     flag,product_name,brand_name,GMV,cat1_name,cat2_name,brand_id,product_id = product_list
    #     with open("./output_brand.txt","a",encoding="utf-8") as f2:
    #         f2.write("\t".join([str(product_id),product_name,str(brand_id),brand_name]) + "\n")
    #         f2.flush()
#
# def check_lack_douyin_v1(input_file):
#     try:
#         conn = hive.connect(host='172.20.207.6', port=10000, username='supdev')
#         # conn = connect(host='172.20.207.6', port=10000, auth_mechanism="PLAIN")
#         cur = conn.cursor()
#         sql_str = 'SELECT brand_name,count(1) AS num from research_online.live_product_maintain_for_brand where brand_match_result = 0 and product_source = "抖音-自营" GROUP BY brand_name ORDER BY num DESC'
#         sql = 'select * from (select *,row_number() over(partition by q.brand_id_std order by new_gmv desc)rn from tmp.gdj_douyin_goods_list_b q)a where rn <= 100'
#         sql = "select q.product_id,q.title,q.brand_words,q.title_pinjie,q.new_gmv,q.brand_id_std,q.brand_name_std,q.category1_id,q.category1,q.match_type,q.goods_url from tmp.gdj_douyin_goods_list_b q where q.match_type not in('没有匹配到任何品牌')"
#         cur.execute(sql)
#         data = cur.fetchall()
#         cur.close()
#         conn.close()
#         return data
#
#     except Exception as e:
#         print(traceback.format_exc())

def check_lack_douyin(input_file,standard_file):
    idx = 0
    product_dict = {}
    B_brand_dict = {}
    A_brand_dict = {}
    sort_A_brand_dict = {}
    sort_B_brand_dict = {}
    brand_stat_dict = {}
    B_top100_product_list = []
    A_B_top100_product_list = []


    brand_dict,ori_brand_dict = split_standard_brand(standard_file)
    print(len(brand_dict))
    with open(input_file,"r",encoding="utf-8") as f1:
        for line in f1:
            idx += 1
            if idx%10000 == 0:print(idx)
            if len(line) == 0:continue
            line_list = line.strip().split("\t")
            if len(line_list) != 14:continue
            line_list = [tmp.strip() for tmp in line_list]
            product_id,product_name,brand_words,key_words,category1_id,category1_name,brand_id,brand_name,match_type,dt,rn,sale_count,gmv,url = line_list
            key_words_ = re.sub('\W+', '', key_words).replace("_", '').lower()
            product_dict[product_id.strip()] = line_list
            if brand_id not in B_brand_dict:
                B_brand_dict[brand_id] = [product_id]
            else:
                B_brand_dict[brand_id].append(product_id)
            for brand_id_item,brand_name_item_list in brand_dict.items():
                for brand_name_item in brand_name_item_list:
                    # brand_name_item_ = re.findall('[\u4e00-\u9fa5a-zA-Z0-9]+',brand_name_item,re.S)
                    # brand_name_item_ = re.sub('\W+', '', brand_name_item).replace("_", '')
                    # if brand_id_item == '10421087':
                    #     print(brand_name_item + "   " + key_words)
                    # key_words_ = re.sub('\W+', '', key_words).replace("_", '')
                    if len(key_words_.split(brand_name_item)) > 1:
                        if brand_id_item not in A_brand_dict:
                            A_brand_dict[brand_id_item] = [product_id]
                        else:
                            A_brand_dict[brand_id_item].append(product_id)
                        break

    for item,values_list in A_brand_dict.items():
        sorted_list = sorted(values_list,key = lambda k:float(product_dict[k][12]),reverse=True)
        sort_A_brand_dict[item] = sorted_list

    for item,values_list in B_brand_dict.items():
        sorted_list = sorted(values_list,key = lambda k:float(product_dict[k][12]),reverse=True)
        sort_B_brand_dict[item] = sorted_list


    for brand in ori_brand_dict:
        brand_name = ori_brand_dict[brand]
        A_gmv = 0.0
        A_product_num = 0
        B_gmv = 0.0
        B_product_num = 0
        B_top100_gmv = 0.0
        A_B_brand_list = []
        A_B_gmv = 0.0
        A_B_top100_gmv = 0.0
        B_divide_A_num = 0.0
        B_divide_A_gmv = 0.0
        B_top100_gmv_rate = 0.0
        A_B_top100_gmv_rate = 0.0

        if brand in sort_A_brand_dict:
            A_product_num = len(sort_A_brand_dict[brand])
            for product_item in sort_A_brand_dict[brand]:
                A_gmv = A_gmv + float(product_dict[product_item][12])
                if brand not in sort_B_brand_dict:
                    A_B_brand_list = sort_A_brand_dict[brand]
                else:
                    if product_item not in sort_B_brand_dict[brand]:
                        A_B_brand_list.append(product_item)

        if brand in sort_B_brand_dict:
            index = 0
            B_product_num = len(sort_B_brand_dict[brand])
            for product_item in sort_B_brand_dict[brand]:
                index += 1
                B_gmv = B_gmv + float(product_dict[product_item][12])
                if index <= 100:
                    B_top100_gmv = B_top100_gmv + float(product_dict[product_item][12])
                    B_top100_product_list.append(str(brand) + "\t" + str(brand_name) + "\t" + str(product_item) + "\t" + str(product_dict[product_item][12]) + "\t" + str(product_dict[product_item][4]) + "\t" + str(product_dict[product_item][5]) + "\t" + str(product_dict[product_item][8]) + "\t" + str(product_dict[product_item][2]) + "\t" + str(product_dict[product_item][1]))


        if len(A_B_brand_list) != 0:
            index = 0
            for product in A_B_brand_list:
                index += 1
                A_B_gmv = A_B_gmv + float(product_dict[product][12])
                if index <= 100:
                    A_B_top100_gmv = A_B_top100_gmv + float(product_dict[product][12])
                    A_B_top100_product_list.append(str(brand) + "\t" + str(brand_name) + "\t" + str(product) + "\t" + str(product_dict[product][12]) + "\t" + str(product_dict[product][6]) + "\t" + str(product_dict[product][7]) + "\t" + str(product_dict[product][4]) + "\t" + str(product_dict[product][5]) + "\t" + str(product_dict[product][8]) + "\t" + str(product_dict[product][2]) + "\t" + str(product_dict[product][1]))


        if A_product_num >= B_product_num and A_product_num != 0:
            B_divide_A_num = B_product_num/A_product_num

        if A_gmv >= B_gmv and A_gmv != 0.0:
            B_divide_A_gmv = B_gmv / A_gmv

        if B_gmv != 0.0:
            B_top100_gmv_rate = B_top100_gmv/B_gmv
        if A_B_gmv != 0.0:
            A_B_top100_gmv_rate = A_B_top100_gmv/A_B_gmv

        brand_stat_dict[brand] = [brand,brand_name,A_product_num,A_gmv,B_product_num,B_gmv,B_top100_gmv,A_B_top100_gmv,B_divide_A_num,B_divide_A_gmv,B_top100_gmv_rate,A_B_top100_gmv_rate]


    return brand_stat_dict,B_top100_product_list,A_B_top100_product_list



def split_standard_brand(standard_file):
    brand_dict = {}
    ori_brand_dict = {}
    with open(standard_file,"r",encoding="utf-8") as f1:
        for line in f1:
            if len(line) == 0:continue
            brand_name = line.strip()
            brand_name_list = brand_name.split("\t")
            if len(brand_name_list) != 2:continue
            brand_name_id,brand_name = brand_name_list
            ori_brand_dict[brand_name_id.strip()] = brand_name.strip()
            name_list = [tmp.lower() for tmp in brand_name.split("/")]
            name_list_ = []
            for item in name_list:
                brand_name_item_ = re.sub('\W+', '', item).replace("_", '')
                name_list_.append(brand_name_item_)
            brand_dict[brand_name_id.strip()] = name_list_
    return brand_dict,ori_brand_dict

def insert():
    #连接数据库
    conn = MongoClient("XXX.XXX.XXX.XXX:XXXX",maxPoolSize=None)
    my_db = conn['test']
    my_collection = my_db['test_1']

    # 批量写
    i = 0
    t0 = time.time()
    data =[]
    while True:
        #'_id'为主键，循环时递增，全部添加到data列表内
        data.append(InsertOne({"_id":i,"insert_time": int(time.time() * 1000)}))
        i+=1
        #判断列表长度，达到10000执行插入，后继续循环
        if len(data) == 10000:
            my_collection.bulk_write(data)
            res = []
            i += 1
            continue
        #判断i等于1亿时停止循环
        elif i == 100000000:
             break


def save_mongo_db():
    MyMongodb = MongoClient(
        'mongodb://spider:spidermining@172.20.207.10:27051,172.20.207.12:27051,172.20.207.13:27051/admin')
    db = MyMongodb.get_database('pdd_map_jd_cate')
    collect = db.get_collection('query_jd_id')

    idx = 0
    with open("./data/spu_id_jd.txt","r",encoding="utf-8") as f1:
        for line in f1:
            if len(line) == 0:continue
            if idx == 0:
                spu_id_list = []
                tmp_dict = {}
            if idx%50000 == 0 and idx > 0:
                collect.insert_many(spu_id_list)
                spu_id_list = []
                tmp_dict = {}
                print("50000")
            pid = line.strip()
            tmp_dict["pid"] = str(pid)
            spu_id_list.append(tmp_dict)
            idx += 1


def list_of_groups(list_info, per_list_len):
    '''
    :param list_info:   列表
    :param per_list_len:  每个小列表的长度
    :return:
    '''
    list_of_group = zip(*(iter(list_info),) *per_list_len)
    end_list = [list(i) for i in list_of_group] # i is a tuple
    count = len(list_info) % per_list_len
    end_list.append(list_info[-count:]) if count !=0 else end_list
    return end_list


def get_jd_data_hive():
    '''
    从hive中取出来的数据类型为list，以及每一项都为一个tuple，每一个单元格的数据都为tuple中的一个元素.
    2235063
    <class 'list'>
    ('62456977355',)
    :return:
    '''
    try:
        idx = 0
        t0 = time.time()
        conn = hive.connect(host='172.20.207.6', port=10000, username='supdev')
        # conn = connect(host='172.20.207.6', port=10000, auth_mechanism="PLAIN")
        cur = conn.cursor()
        sql = "select spu_id from dim.dim_retailers_online_spu_sku where platform_type = 'jd'"
        cur.execute(sql)
        data = cur.fetchall()
        print("finish getting data!")
        ret_data = list_of_groups(data,24000000)
        print("hive get data finish!")
        for ret_item in ret_data:
            idx += 1
            with open("./data/output_data_" + str(idx) + ".txt","w",encoding="utf-8") as f1:
                for item in ret_item:
                    f1.write(item[0] + "\n")
                f1.flush()

        cur.close()
        conn.close()
        t1 = time.time()
        return data

    except Exception as e:
        print(traceback.format_exc())


def meizhuang_brand_choose(input_file):
    pass


if __name__ == "__main__":
    idx = 0
    #list取数据
    # data_result = mysql_opt_v1()
    # print(len(data_result))
    # data_result_list = list(data_result)
    # print(data_result_list)
    # for brand_item in data_result_list:
    #     for data_item in brand_item:
    #         idx += 1
    #         print(str(data_item[0]) + "\t" + str(data_item[1]) + "\t" + str(data_item[2] + "\t" + str(data_item[3])  + "\t" + str(data_item[4])))
    # print(idx)
    # with open("./data/check_not.txt","w",encoding="utf-8") as f1:
    #     for item in data_result_list:
    #         f1.write(str(item[0]) + "\t" + str(item[1]) + "\t" + str(item[2]) + "\t" + str(item[3]) + "\t" + str(item[4]) + "\t" + str(item[5]) +  "\n")
    #     f1.flush()

    # # data取数据
    # data = stat_brand_num()
    # for brand_item in data:
    #     idx += 1
    #     print(str(brand_item[0]) + "\t" + str(brand_item[1]))
    # print(idx)

    #筛选未审核上线品牌
    # get_online_brand_v1()

    #通过抖音品牌获取商品信息
    # data_result = get_product()
    # print(len(data_result))
    # data_result_list = list(data_result)
    # print(data_result_list)
    # for brand_item in data_result_list:
    #     for data_item in brand_item:
    #         idx += 1
    #         print(str(data_item[0]) + "\t" + str(data_item[1]) + "\t" + str(data_item[2] + "\t" + str(data_item[3])  + "\t" + str(data_item[4])))
    # print(idx)

    #收集所有文件信息
    # collect_data()

    #比较两个文件品牌的交集
    # compare_brand()

    #查漏环节开发
    brand_file = "./z_douyin_clean_pro/xiaowu_standard_brand.txt"
    input_file = "./douyin_lack_data.txt"
    brand_stat_dict,B_top100_product_list,A_B_top100_product_list = check_lack_douyin(input_file, brand_file)
    with open("./data/check_brand_stat.txt","w",encoding="utf-8") as f1:
        for brand,vlaues_list in brand_stat_dict.items():
            vlaues_list = [str(tmp) for tmp in vlaues_list]
            f1.write("\t".join(vlaues_list))
            f1.write("\n")
        f1.flush()
    with open("./data/B_top100_product.txt","w",encoding="utf-8") as f2:
        f2.write("\n".join(B_top100_product_list))
    with open("./data/A_B_top100_product.txt", "w", encoding="utf-8") as f3:
        f3.write("\n".join(A_B_top100_product_list))

    #将spu_id插入到Mongo库中
    # save_mongo_db()

    #从hive中取到京东数据
    # data = get_jd_data_hive()
    # print(len(data))