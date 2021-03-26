#!/usr/bin/env python3
#coding=utf-8
"""
一、数据来源：
    分母：某一个品牌的全量gmv来自2020-10-21日识别的数据
        1)来自殿军的hive表(已经有sql语句了)
    分子：
        1)来自实习生的审核的品牌数据。这个数据仅仅是一部分当时情况的高gmv的商品
        2)数据来自包骏线上品牌审核mysql数据库
二、输出数据的格式
    1)格式：品牌编号、品牌名称、全部商品个数、 全部商品gmv、已审核的商品个数、已审核商品gmv、已审核商品数量占比、已审核商品gmv占比
    2)gmv占比降序排列
    3)单个品牌商品gmv的降序排列
三、缺点明显，但具有参考价值
"""

import sys, os
import pymysql
import pyhive
import datetime
import traceback
from openpyxl import Workbook
rate_threshold = 0.8

t_lst1 = ('品牌编号', '品牌名称', '全部商品个数', '全部商品gmv', '已审核的商品个数', '已审核商品gmv', '已审核商品数量占比', '已审核商品gmv占比')


def mysql_opt():
    try:
        db = pymysql.connect(host='rm-8vb1dlcr1x89lo03z.mysql.zhangbei.rds.aliyuncs.com', \
                             port=3306, user='edw', \
                             passwd='0Zwafqams%GDL3BU', db='mysql', \
                             charset='utf8')

        cursor = db.cursor()
        cursor.execute("select product_id, brand_id, brand_name from research_online.live_product_maintain_for_brand q where brand_match_result > 0")
        data = cursor.fetchall()
        cursor.close()
        db.close()
        return data
    except:
        print(traceback.format_exc())

def saving_mysql_data():
    r_lst = []
    b_dict = {}
    data = mysql_opt()
    for d in data:
        p_id, b_id, b_name = d
        b_dict[b_id] = ''
        r_lst.append("%s\t%s\t%s\t" % d)

    print("已经审核的品牌数量:", len(b_dict))
    with open("gmv_output/checked_product.txt", "w",encoding="utf-8") as f2:
        f2.write("\n".join(r_lst))
        f2.flush()

#saving_mysql_data()

def wrt_one_sheet(wb, ori_sheet_name, lst8):
    try:
        sheet_name = ori_sheet_name.strip().replace("/", "_").replace("、", "")
        wb.create_sheet(sheet_name, 0)
        sh_obj = wb[sheet_name]

        for m in range(1, len(lst8) + 1):
            lst7 = lst8[m - 1]
            for n in range(1, len(lst7) + 1):
                sh_obj.cell(row=m, column=n).value = lst7[n - 1]

        return wb
    except:
        print(traceback.format_exc())
        print(ori_sheet_name)

def wrt_brand_excel(lst8, save_file_name):
    wb = Workbook()
    '''
    wb.create_sheet("审核品牌", 0)
    sh_obj = wb["审核品牌"]
    
    for j in range(1, len(t_lst1) + 1):
        sh_obj.cell(row=1, column=j).value = t_lst1[j - 1]
    for m in range(1, len(lst8) + 1):
        lst7 = lst8[m - 1]
        for n in range(1, len(lst7) + 1):
            sh_obj.cell(row=m + 1, column=n).value = str(lst7[n - 1])
    '''
    wrt_one_sheet(wb, "审核品牌", lst8)
    if os.path.exists(save_file_name):
        os.remove(save_file_name)
    wb.save(save_file_name)

def wrt_cat1_excel(cat1_stat_dict, save_file_name):
    wb = Workbook()
    for k,v in cat1_stat_dict.items():
        wrt_one_sheet(wb, k, v)

    if os.path.exists(save_file_name):
        os.remove(save_file_name)
    wb.save(save_file_name)

def wrt_whole_report_excel(overview_lst, brand_all_excel_lst, brand_lower_excel_lst, cat1_excel_dict):
    wb = Workbook()
    for k, v in cat1_excel_dict.items():
        wrt_one_sheet(wb, k, v)
    wrt_one_sheet(wb, "小于80-percent品牌", brand_lower_excel_lst)
    wrt_one_sheet(wb, "全部品牌", brand_all_excel_lst)
    wrt_one_sheet(wb, "测评结果", overview_lst)
    f_name = "gmv_output/抖音品牌审核报表_%s.xlsx" % datetime.datetime.now().strftime('%Y-%m-%d-%H')
    if os.path.exists(f_name):
        os.remove(f_name)
    wb.save(f_name)

class GMVEvalFunc(object):
    def __init__(self, all_product_gmv_file, standard_brand_file):
        if not os.path.exists(all_product_gmv_file):
            raise Exception("%s 文件不存在!" % all_product_gmv_file)

        if not os.path.exists(standard_brand_file):
            raise Exception("%s 文件不存在!" % standard_brand_file)
        self.all_product_gmv_file = all_product_gmv_file
        self.standard_brand_file = standard_brand_file

        self.checked_brand_product_dict, self.checked_brand_dict = \
            self.getting_checked_product_dict()
        self.cat1_brand_dict = self.getting_cat1_brand_info()
        self.all_brand_product_dict, self.all_product_gmv_dict, \
        self.all_product_gmv_dict_ext, self.all_gmv = self.loading_all_product_info()

        self.product_cat1_dict = self.getting_product_cat1_info()
        self.brand_gmv = 0.0
        self.checked_brand_gmv = 0.0
        self.qualified_checked_brand_gmv = 0.0

        self.checked_right_product_brand_dict = {} # {b_id|b_name: [p_id, p_id]}
        pass

    def getting_checked_product_dict(self):
        b_id_dict = {}
        with open("../../z_douyin_clean_pro_/xiaowu_standard_brand.txt", "r", encoding="utf-8") as f1:
            for line in f1:
                line = line.strip()
                if line == "": continue
                # brand_id, brand_name, cat1_id, cat1, gmv
                lst1 = line.split("\t")
                if len(lst1) != 2:
                    continue
                lst1 = [tmp.strip() for tmp in lst1]
                b_id, b_name = lst1
                b_id_dict[b_id] = ''

        data =  mysql_opt()
        checked_brand_dict = {}
        checked_brand_product_dict = {} # {'b_id|b_name': [p_id, p_id]}
        for d in data:
            p_id, b_id, b_name = d
            if p_id == None or b_id == None or b_name == None:
                print(p_id)
                continue
            b_id = str(b_id).strip()
            if b_id not in b_id_dict: continue
            p_id = str(p_id).strip()
            b_name = b_name.strip()
            k1 = "%s|%s" % (b_id, b_name)
            checked_brand_dict[k1] = ''
            if k1 in checked_brand_product_dict:
                z = checked_brand_product_dict[k1]
                z = list(set(z + [p_id]))
                checked_brand_product_dict[k1] = z
            else:
                checked_brand_product_dict[k1] = [p_id]

        return checked_brand_product_dict, checked_brand_dict

    def getting_cat1_brand_info(self):
        cat1_brand_dict = {}
        with open(self.standard_brand_file, "r", encoding="utf-8") as f1:
            for line in f1:
                line = line.strip()
                if line == "": continue
                lst1 = line.split("\t")
                if len(lst1) != 5:
                    continue
                lst1 = [tmp.strip() for tmp in lst1]
                b_id, ori_b_name, cat1_id, cat1, gmv = lst1
                k1 = "%s|%s" % (b_id, ori_b_name)
                if k1 not in self.checked_brand_dict: continue
                if cat1 in cat1_brand_dict:
                    z = cat1_brand_dict[cat1]
                    y = list(set(z + [k1]))
                    cat1_brand_dict[cat1] = y
                else:
                    cat1_brand_dict[cat1] = [k1]

        return cat1_brand_dict

    def getting_product_cat1_info(self):
        product_cat1_dict = {}
        with open("/douyin_data.txt", "r", encoding="utf-8") as f1:
            for line in f1:
                line = line.strip()
                if line == "": continue
                lst1 = line.split("\001")
                if len(lst1) != 5:
                    continue
                lst1 = [tmp.strip() for tmp in lst1]
                product_id, product_name, brand_word, cat1_id, cat1 = lst1
                if product_id not in self.all_product_gmv_dict: continue
                product_cat1_dict[product_id] = cat1

        return product_cat1_dict

    def loading_all_product_info(self):
        t = 0
        all_gmv = 0.0
        all_brand_product_dict = {} # {'b_id|b_name': [p_id, p_id]}
        all_product_gmv_dict = {}
        all_product_gmv_dict_ext = {}
        with open(self.all_product_gmv_file,"r",encoding="utf-8") as f1:
            for line in f1:
                line = line.strip()
                if line == "": continue
                lst1 = line.split('\t')
                if len(lst1) != 7: continue
                lst1 = [tmp.strip() for tmp in lst1]
                t += 1
                # if t % 5000 == 0: print("t:", t)
                # 3402180027381617879     10407445        Gree/格力       格力空调大2匹风无界     格力(GREE)      0.0     88
                p_id, b_id, b_name, p_name, dy_brand, gmv, _ = lst1
                if p_id == "": continue
                k1 = "%s|%s" % (b_id, b_name)
                if k1 not in self.checked_brand_dict: continue
                try:
                    gmv = float(gmv) / 10000
                except:
                    continue
                all_gmv += gmv
                if k1 in all_brand_product_dict:
                    z = all_brand_product_dict[k1]
                    z = list(set(z + [p_id]))
                    all_brand_product_dict[k1] = z
                else:
                    all_brand_product_dict[k1] = [p_id]

                all_product_gmv_dict[p_id] = gmv
                all_product_gmv_dict_ext[p_id] = [p_name, gmv]

        return all_brand_product_dict, all_product_gmv_dict, all_product_gmv_dict_ext, all_gmv

    def brand_stat(self, brand_info_dict):
        chk_brand_gmv_dict = {}
        for k1, v1 in brand_info_dict.items():
            tmp_gmv = 0.0
            for x in v1:
                if x not in self.all_product_gmv_dict:
                    if k1 in self.checked_right_product_brand_dict:
                        zz = self.checked_right_product_brand_dict[k1]
                        zz += [x]
                        self.checked_right_product_brand_dict[k1] = list(set(zz))
                    else:
                        self.checked_right_product_brand_dict[k1] = [x]

                    print("error1:", k1, x)
                    continue
                tmp_gmv += self.all_product_gmv_dict[x]
            chk_brand_gmv_dict[k1] = (len(v1), tmp_gmv)

        return chk_brand_gmv_dict

    def checked_brand_gmv_rate(self, chked_brand_dict, all_brand_dict):
        for k1, v1 in chked_brand_dict.items():
            _, gmv1 = v1
            self.checked_brand_gmv += gmv1

        for k2, v2 in all_brand_dict.items():
            _, gmv2 = v2
            self.brand_gmv += gmv2

    def brand_stat_info(self):
        chked_brand_dict = self.brand_stat(self.checked_brand_product_dict)
        all_brand_dict = self.brand_stat(self.all_brand_product_dict)
        #
        self.checked_brand_gmv_rate(chked_brand_dict, all_brand_dict)

        all_excel_lst = []
        lower_excel_rate_lst = []
        brand_stat_dict = {}
        for k2, chk_item in chked_brand_dict.items():
            if k2 not in all_brand_dict:
                print("error2:", k2, chk_item)
                continue
            all_item = all_brand_dict[k2]
            chk_num, chk_gmv = chk_item
            all_num, all_gmv = all_item
            num_rate = self.cmp_rate(chk_num, all_num)
            gmv_rate = self.cmp_rate(chk_gmv, all_gmv)
            b_id, b_name = k2.strip().split("|")
            itm= (b_id, b_name, all_num, all_gmv, chk_num, chk_gmv, num_rate, gmv_rate)
            if gmv_rate >= 0.8:
                self.qualified_checked_brand_gmv += chk_gmv
            brand_stat_dict[k2] = itm
            # 品牌编号、品牌名称、全部商品个数、 全部商品gmv、已审核的商品个数、已审核商品gmv、已审核商品数量占比、已审核商品gmv占比
            all_excel_lst.append(itm)
            if gmv_rate < 0.8:
                lower_excel_rate_lst.append(itm)

        all_excel_lst = sorted(all_excel_lst, key=lambda x: x[7], reverse=True)
        all_excel_lst = [t_lst1] + all_excel_lst
        # wrt_brand_excel(all_excel_lst, "gmv_output/已审核品牌状况.xlsx")

        lower_excel_rate_lst = sorted(lower_excel_rate_lst, key=lambda x: x[7], reverse=True)
        lower_excel_rate_lst = [t_lst1] + lower_excel_rate_lst
        #wrt_brand_excel(lower_excel_rate_lst, "gmv_output/GMV占比低于80-percent.xlsx")

        return brand_stat_dict, all_excel_lst, lower_excel_rate_lst

    def cmp_rate(self,fz, fm):
        if fm != 0:
            r = round(1.0 * fz / fm, 4)
        else:
            r = 0.0
        if r > 1.0: r = 1.0
        return r

    def cat1_stat_info(self, brand_stat_dict):
        cat1_stat_dict = {}
        cat1_stat_lst = []
        for k, v in self.cat1_brand_dict.items():
            if len(v) == 0: continue
            t_all_num, t_all_gmv, t_chk_num, t_chk_gmv = 0.0, 0.0, 0.0, 0.0
            tmp_lst = []
            for z in v:
                if z not in brand_stat_dict: continue
                item = brand_stat_dict[z]
                b_id, b_name, all_num, all_gmv, chk_num, chk_gmv, num_rate, gmv_rate = item
                t_all_num += all_num
                t_all_gmv += all_gmv
                t_chk_num += chk_num
                t_chk_gmv += chk_gmv
                tmp_lst.append(item)
            tmp_lst = sorted(tmp_lst, key=lambda x: x[7], reverse=True)
            cat1_num_rate = self.cmp_rate(t_chk_num, t_all_num)
            cat1_gmv_rate = self.cmp_rate(t_chk_gmv, t_all_gmv)
            cat1_info = ("一级类目总体状况:", k, t_all_num, t_all_gmv, t_chk_num, t_chk_gmv, cat1_num_rate, cat1_gmv_rate)
            ok_lst = [cat1_info, ('', '','','','','','',''), ('', '','','','','','',''), t_lst1]
            ok_lst += tmp_lst
            cat1_stat_dict[k] = ok_lst
            cat1_stat_lst.append(cat1_info)

        '''
        cat1_stat_lst = sorted(cat1_stat_lst, key=lambda x: x[7], reverse=True)
        for tmp in cat1_stat_lst:
            print("%s\t%s\t%s\t%s\t%s\t%s\t%s\t%s" % tmp)

        print()
        print()
        print()
        cat1_stat_lst = sorted(cat1_stat_lst, key=lambda x: x[3], reverse=True)
        for tmp in cat1_stat_lst:
            print("%s\t%s\t%s" % (tmp[1], tmp[3], round(tmp[3]/self.all_gmv, 5)))
        '''

        #wrt_cat1_excel(cat1_stat_dict, "gmv_output/一级类目审核.xlsx")

        return cat1_stat_dict

    def stat_go(self):
        brand_stat_dict, brand_all_excel_lst, brand_lower_excel_lst = self.brand_stat_info()

        cat1_excel_dict = self.cat1_stat_info(brand_stat_dict)


        print("全部品牌gmv: ", self.brand_gmv)
        print("已审核品牌gmv: ", self.checked_brand_gmv)
        print("已审核品牌gmv占比: ", round(self.checked_brand_gmv / self.brand_gmv, 5))
        print("gmv占比大于80%已审核品牌gmv: ", self.qualified_checked_brand_gmv, )
        print("gmv占比大于80%已审核品牌gmv占比: ", round(self.qualified_checked_brand_gmv / self.brand_gmv, 5))

        lst1 = [
            ("全部品牌gmv:", self.brand_gmv),
            ("已审核品牌gmv:", self.checked_brand_gmv),
            ("已审核品牌gmv占比:", round(self.checked_brand_gmv / self.brand_gmv, 5)),
            ("gmv占比大于80%已审核品牌gmv: ", self.qualified_checked_brand_gmv),
            ("gmv占比大于80%已审核品牌gmv占比:", round(self.qualified_checked_brand_gmv / self.brand_gmv, 5))
        ]
        wrt_whole_report_excel(lst1, brand_all_excel_lst, brand_lower_excel_lst, cat1_excel_dict)

        #print(self.checked_right_product_brand_dict)

    def xiaowu_brand_top10_product_stat(self):
        # 小米  p_id,p_name, gmv_top-10
        chked_brand_dict = self.brand_stat(self.checked_brand_product_dict)
        # b_id|b_name p_num, brand_gmv
        b_lst_sorted = [(k, v[0], v[1]) for k,v in chked_brand_dict.items()]
        b_lst_sorted = sorted(b_lst_sorted, key=lambda x:x[2], reverse=True)
        r_lst = []
        for tmp in b_lst_sorted:
            k1 = tmp[0]
            if k1 not in self.checked_brand_product_dict: continue
            b_id, b_name = k1.strip().split('|')
            v = self.checked_brand_product_dict[k1]
            tmp_lst = []
            for p_id in v:
                if p_id not in self.all_product_gmv_dict_ext: continue
                p_name, p_gmv = self.all_product_gmv_dict_ext[p_id]
                tmp_lst.append((p_id, p_name, p_gmv))
            tmp_lst = sorted(tmp_lst, key=lambda x: x[2], reverse=True)
            for z in tmp_lst[:10]:
                r_lst.append("%s\t%s\t%s\t%s\t%s" % (b_id, b_name, z[0], z[1], z[2]))
            r_lst.append("%s\t%s\t%s\t%s\t%s" % ('', '', '', '', ''))
            r_lst.append("%s\t%s\t%s\t%s\t%s" % ('', '', '', '', ''))

        with open("gmv_output/品牌-高GMV-Top10商品.txt", "w",encoding="utf-8") as f1:
            f1.write("\n".join(r_lst))
            f1.flush()

    def xiaowu_brand_cat1_gmv(self):
        # 以下数据都为已审核的品牌的数据
        # 小米 小米_gmv 商品数, cat1_name cat1_gmv cat1_num,......
        # 小米  p_id,p_name, gmv_top-10
        chked_brand_dict = self.brand_stat(self.checked_brand_product_dict)
        # b_id|b_name p_num, brand_gmv
        b_lst_sorted = [(k, v[0], v[1]) for k, v in chked_brand_dict.items()]
        b_lst_sorted = sorted(b_lst_sorted, key=lambda x: x[2], reverse=True)
        r_lst = []
        for tmp in b_lst_sorted:
            k1 = tmp[0]
            if k1 not in self.checked_brand_product_dict: continue
            b_id, b_name = k1.strip().split('|')
            v = self.checked_brand_product_dict[k1]
            tmp_dict = {}
            tmp_dict_2 = {}
            for p_id in v:
                if p_id not in self.product_cat1_dict: continue
                if p_id not in self.all_product_gmv_dict: continue
                cat1 = self.product_cat1_dict[p_id]
                gmv = self.all_product_gmv_dict[p_id]
                if cat1 in tmp_dict:
                    x = tmp_dict[cat1]
                    tmp_dict[cat1] = x + gmv
                else:
                    tmp_dict[cat1] = gmv

                if cat1 in tmp_dict_2:
                    xx = tmp_dict_2[cat1]
                    tmp_dict_2[cat1] = xx + 1
                else:
                    tmp_dict_2[cat1] = 1

            lst9 = [(k2, v2) for k2, v2 in tmp_dict.items()]
            lst9 = sorted(lst9, key=lambda x: x[1], reverse=True)
            lst10 = []
            b_num = 0
            b_gmv = 0.0
            for y in lst9:
                cat1, cat1_gmv = y
                cat1_num = tmp_dict_2[cat1]
                b_num += cat1_num
                b_gmv += cat1_gmv
                lst10.append("%s\t%s\t%s" % (cat1, cat1_num, cat1_gmv))

            r_lst.append("%s\t%s\t%s\t%s" %(b_name, b_num, b_gmv, "\t".join(lst10)))

        with open("gmv_output/品牌-一级类目-GMV-统计.txt", "w", encoding="utf-8") as f5:
            f5.write("\n".join(r_lst))
            f5.flush()

    def sencond_hand_gmv_stat(self):
        tmp_gmv = 0.0
        for pid, gmv in self.all_product_gmv_dict.items():
            if pid not in self.product_cat1_dict: continue
            cat1 = self.product_cat1_dict[pid]
            if cat1 != "二手商品": continue

            tmp_gmv += gmv

        print("二手商品gmv:", tmp_gmv, "占比:", round(tmp_gmv/self.all_gmv, 5))

if __name__ == "__main__":
    obj = GMVEvalFunc("z_gmv.txt", "C:/Users/Cwgong/PycharmProjects/z_douyin_clean_pro_update/standard_brand_info.txt")
    obj.brand_stat_info()
    #obj.stat_go()
    #obj.xiaowu_brand_top10_product_stat()
    obj.xiaowu_brand_cat1_gmv()
    #obj.sencond_hand_gmv_stat()